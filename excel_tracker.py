"""
excel_tracker.py
----------------
Exporte les propositions du bot (live + back-test) dans un Excel complet.
Feuilles : Dashboard, Value Bets Live, Back-test, P&L par saison.

Usage :
  python excel_tracker.py                    → génère tracker vide
  python excel_tracker.py --backtest         → inclut les résultats du back-test
  python excel_tracker.py --add-bet          → ajoute un pari manuellement
"""

import json
import argparse
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "nba_value_tracker.xlsx"

# ── Couleurs ──────────────────────────────────────────────────────────────────
C_DARK    = "1A1A2E"   # fond sombre
C_BLUE    = "0F3460"   # bleu NBA
C_ACCENT  = "E94560"   # rouge/rose accent
C_GREEN   = "16C79A"   # vert gain
C_YELLOW  = "F5A623"   # jaune warning
C_WHITE   = "FFFFFF"
C_LGRAY   = "F2F2F2"
C_GRAY    = "CCCCCC"
C_WIN     = "C6EFCE"   # fond vert clair (gagné)
C_LOSS    = "FFC7CE"   # fond rouge clair (perdu)
C_PENDING = "FFEB9C"   # fond jaune (en attente)


def style_header(cell, bg=C_BLUE, fg=C_WHITE, bold=True, size=11):
    cell.font         = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill         = PatternFill("solid", start_color=bg)
    cell.alignment    = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_title(cell, text, bg=C_DARK, fg=C_WHITE, size=14):
    cell.value        = text
    cell.font         = Font(bold=True, color=fg, size=size, name="Arial")
    cell.fill         = PatternFill("solid", start_color=bg)
    cell.alignment    = Alignment(horizontal="left", vertical="center")


def thin_border():
    s = Side(style="thin", color=C_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)


def pct(val):
    return f"{val*100:.1f}%" if val is not None else "-"


# ── Feuille Dashboard ─────────────────────────────────────────────────────────

def create_dashboard(wb: Workbook, live_bets: list, backtest: list):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    # Titre
    ws.merge_cells("B1:F1")
    style_title(ws["B1"], "NBA VALUE BOT - DASHBOARD", C_DARK, C_WHITE, 16)
    ws.row_dimensions[1].height = 35

    ws.merge_cells("B2:F2")
    ws["B2"].value     = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws["B2"].font      = Font(color="888888", size=9, name="Arial")
    ws["B2"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 18

    # ── KPIs back-test ────────────────────────────────────────────────────────
    bets     = [b for b in backtest if b.get("bet_placed")]
    wins     = [b for b in bets if b.get("won")]
    pnl_bt   = sum(b.get("pnl", 0) for b in bets)
    invested = sum(b.get("kelly_stake", 0) for b in bets)
    roi_bt   = pnl_bt / invested * 100 if invested else 0
    wr_bt    = len(wins) / len(bets) * 100 if bets else 0

    kpis = [
        ("Paris back-test",  len(bets),          C_BLUE),
        ("Taux de victoire", f"{wr_bt:.1f}%",     C_GREEN if wr_bt > 52 else C_ACCENT),
        ("P&L théorique",    f"{pnl_bt:+.0f}€",  C_GREEN if pnl_bt > 0 else C_ACCENT),
        ("ROI",              f"{roi_bt:.1f}%",    C_GREEN if roi_bt > 0 else C_ACCENT),
        ("Paris live",       len(live_bets),       C_BLUE),
    ]

    ws.row_dimensions[4].height = 18
    ws.merge_cells("B3:F3")
    ws["B3"].value = "PERFORMANCE BACK-TEST"
    ws["B3"].font  = Font(bold=True, color=C_BLUE, size=11, name="Arial")

    for i, (label, value, color) in enumerate(kpis):
        col = get_column_letter(2 + i)
        ws.row_dimensions[5].height = 55
        ws.row_dimensions[6].height = 25
        cell_val   = ws[f"{col}5"]
        cell_label = ws[f"{col}6"]

        cell_val.value     = value
        cell_val.font      = Font(bold=True, color=C_WHITE, size=18, name="Arial")
        cell_val.fill      = PatternFill("solid", start_color=color)
        cell_val.alignment = Alignment(horizontal="center", vertical="center")

        cell_label.value     = label
        cell_label.font      = Font(color="555555", size=9, name="Arial")
        cell_label.fill      = PatternFill("solid", start_color=C_LGRAY)
        cell_label.alignment = Alignment(horizontal="center")

    # ── P&L par saison ────────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 20
    ws.merge_cells("B8:F8")
    ws["B8"].value = "P&L PAR SAISON"
    ws["B8"].font  = Font(bold=True, color=C_BLUE, size=11, name="Arial")

    seasons_data = {}
    for b in bets:
        s = b.get("season", "?")
        if s not in seasons_data:
            seasons_data[s] = {"bets": 0, "wins": 0, "pnl": 0.0, "invested": 0.0}
        seasons_data[s]["bets"]     += 1
        seasons_data[s]["wins"]     += 1 if b.get("won") else 0
        seasons_data[s]["pnl"]      += b.get("pnl", 0)
        seasons_data[s]["invested"] += b.get("kelly_stake", 0)

    headers = ["Saison", "Paris", "Victoires", "Taux V.", "P&L (€)", "ROI"]
    for j, h in enumerate(headers):
        c = ws.cell(row=9, column=2+j, value=h)
        style_header(c, C_BLUE)
        ws.row_dimensions[9].height = 22

    for i, (season, d) in enumerate(sorted(seasons_data.items())):
        row  = 10 + i
        wr   = d["wins"] / d["bets"] * 100 if d["bets"] else 0
        roi  = d["pnl"] / d["invested"] * 100 if d["invested"] else 0
        vals = [season, d["bets"], d["wins"], f"{wr:.1f}%",
                f"{d['pnl']:+.2f}", f"{roi:.1f}%"]
        bg   = C_WIN if d["pnl"] > 0 else C_LOSS
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=2+j, value=v)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center")
            c.border    = thin_border()
        ws.row_dimensions[row].height = 20

    # ── Derniers paris live ───────────────────────────────────────────────────
    start_row = 10 + len(seasons_data) + 2
    ws.merge_cells(f"B{start_row}:F{start_row}")
    ws[f"B{start_row}"].value = "DERNIERS PARIS LIVE"
    ws[f"B{start_row}"].font  = Font(bold=True, color=C_BLUE, size=11, name="Arial")

    live_headers = ["Date", "Match", "Marché", "Cote", "Value", "Mise", "Résultat"]
    for j, h in enumerate(live_headers):
        c = ws.cell(row=start_row+1, column=2+j, value=h)
        style_header(c, C_ACCENT)

    if not live_bets:
        ws.cell(row=start_row+2, column=2,
                value="Aucun pari live enregistré — lance le bot !").font = Font(
                    color="888888", italic=True, name="Arial")
    else:
        for i, bet in enumerate(live_bets[-10:]):
            row = start_row + 2 + i
            result = bet.get("result", "⏳")
            bg = C_WIN if result == "WON" else (
                 C_LOSS if result == "LOST" else C_PENDING)
            vals = [
                bet.get("date", ""),
                f"{bet.get('away_team','')} @ {bet.get('home_team','')}",
                bet.get("market", ""),
                bet.get("bookie_odd", ""),
                pct(bet.get("value")),
                f"{bet.get('kelly_stake', 0):.0f}€",
                result,
            ]
            for j, v in enumerate(vals):
                c = ws.cell(row=row, column=2+j, value=v)
                c.font      = Font(name="Arial", size=10)
                c.fill      = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(horizontal="center")
                c.border    = thin_border()


# ── Feuille Value Bets Live ───────────────────────────────────────────────────

def create_live_sheet(wb: Workbook, live_bets: list):
    ws = wb.create_sheet("Value Bets Live")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 12

    ws.merge_cells("A1:J1")
    style_title(ws["A1"], "VALUE BETS LIVE - Suivi des propositions du bot")
    ws.row_dimensions[1].height = 30

    headers = ["Date", "Match", "Marché", "Cote", "Total prédit",
               "Ligne bookie", "Value", "P(gagner)", "Mise (€)", "Résultat"]
    for j, h in enumerate(headers):
        c = ws.cell(row=2, column=1+j, value=h)
        style_header(c, C_BLUE)
    ws.row_dimensions[2].height = 22

    for i, bet in enumerate(live_bets):
        row    = 3 + i
        result = bet.get("result", "⏳ En attente")
        bg     = (C_WIN     if result == "WON"  else
                  C_LOSS    if result == "LOST" else
                  C_PENDING)

        vals = [
            bet.get("date", ""),
            f"{bet.get('away_team','')} @ {bet.get('home_team','')}",
            bet.get("market", ""),
            bet.get("bookie_odd", ""),
            bet.get("predicted_total", ""),
            bet.get("total_line", ""),
            pct(bet.get("value")),
            pct(bet.get("model_prob")),
            bet.get("kelly_stake", 0),
            result,
        ]
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=1+j, value=v)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center")
            c.border    = thin_border()
        ws.row_dimensions[row].height = 18

    # Ligne d'ajout manuelle
    next_row = 3 + len(live_bets)
    ws.cell(row=next_row, column=1,
            value="→ Ajouter un pari ici").font = Font(
                color=C_ACCENT, italic=True, bold=True, name="Arial")


# ── Feuille Back-test ─────────────────────────────────────────────────────────

def create_backtest_sheet(wb: Workbook, backtest: list):
    ws = wb.create_sheet("📈 Back-test")
    ws.sheet_view.showGridLines = False

    cols_w = [12, 28, 14, 14, 12, 10, 12, 12, 10, 10, 12]
    for i, w in enumerate(cols_w):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(cols_w))}1")
    style_title(ws["A1"], "📈 BACK-TEST — Résultats historiques 3 saisons")
    ws.row_dimensions[1].height = 30

    headers = ["Date", "Match", "Saison", "Marché", "Total prédit",
               "Ligne", "Réel", "Value", "P(gagner)", "Mise (€)", "P&L (€)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=2, column=1+j, value=h)
        style_header(c, C_DARK)
    ws.row_dimensions[2].height = 22

    bets = [b for b in backtest if b.get("bet_placed")]

    for i, bet in enumerate(bets):
        row    = 3 + i
        won    = bet.get("won", False)
        pnl    = bet.get("pnl", 0)
        bg     = C_WIN if won else C_LOSS

        vals = [
            bet.get("date", ""),
            f"{bet.get('away_team','')} @ {bet.get('home_team','')}",
            bet.get("season", ""),
            bet.get("market", ""),
            bet.get("predicted_total", ""),
            bet.get("line", ""),
            bet.get("actual_total", ""),
            pct(bet.get("value")),
            pct(bet.get("model_prob")),
            bet.get("kelly_stake", 0),
            f"{pnl:+.2f}",
        ]
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=1+j, value=v)
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center")
            c.border    = thin_border()
        ws.row_dimensions[row].height = 16


# ── Feuille P&L cumulé ────────────────────────────────────────────────────────

def create_pnl_sheet(wb: Workbook, backtest: list):
    ws = wb.create_sheet("P&L Cumule")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    style_title(ws["A1"], "P&L CUMULE - Evolution de la bankroll simulee")
    ws.row_dimensions[1].height = 30

    headers = ["#", "Date", "Match", "P&L (€)", "P&L Cumulé (€)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=2, column=1+j, value=h)
        style_header(c, C_GREEN if j >= 3 else C_BLUE)
    ws.row_dimensions[2].height = 22

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    bets       = [b for b in backtest if b.get("bet_placed")]
    cumulative = 0.0

    for i, bet in enumerate(bets):
        row       = 3 + i
        pnl       = bet.get("pnl", 0)
        cumulative += pnl
        bg        = C_WIN if pnl >= 0 else C_LOSS

        vals = [
            i + 1,
            bet.get("date", ""),
            f"{bet.get('away_team','')} @ {bet.get('home_team','')}",
            round(pnl, 2),
            round(cumulative, 2),
        ]
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=1+j, value=v)
            c.font      = Font(name="Arial", size=9,
                               bold=(j >= 3),
                               color="006100" if (j >= 3 and v >= 0) else (
                                     "9C0006" if j >= 3 else "000000"))
            c.fill      = PatternFill("solid", start_color=bg if j == 3 else C_LGRAY)
            c.alignment = Alignment(horizontal="center")
            c.border    = thin_border()
        ws.row_dimensions[row].height = 16

    # Graphique P&L cumulé
    if len(bets) > 1:
        chart = LineChart()
        chart.title  = "Évolution P&L Cumulé (€)"
        chart.style  = 10
        chart.y_axis.title = "P&L Cumulé (€)"
        chart.x_axis.title = "Nombre de paris"
        chart.height = 14
        chart.width  = 26

        data = Reference(ws, min_col=5, min_row=2, max_row=2+len(bets))
        chart.add_data(data, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = C_GREEN
        chart.series[0].graphicalProperties.line.width     = 20000

        ws.add_chart(chart, f"G3")


# ── Main ──────────────────────────────────────────────────────────────────────

def load_live_bets() -> list:
    """Charge les paris live depuis value_bets_log.json."""
    if not os.path.exists("value_bets_log.json"):
        return []
    bets = []
    with open("value_bets_log.json", "r", encoding="utf-8") as f:
        for line in f:
            try:
                entry = json.loads(line.strip())
                for vb in entry.get("value_bets", []):
                    if "result" not in vb:
                        vb["result"] = "⏳ En attente"
                    bets.append(vb)
            except Exception:
                continue
    return bets


def load_backtest() -> list:
    """Charge les résultats du back-test depuis backtest_results.json."""
    if not os.path.exists("backtest_results.json"):
        return []
    with open("backtest_results.json", "r", encoding="utf-8") as f:
        return json.load(f)


def generate_excel(include_backtest: bool = True):
    print(f"[Excel] Chargement des données...")
    live_bets = load_live_bets()
    backtest  = load_backtest() if include_backtest else []

    print(f"[Excel] {len(live_bets)} paris live, {len([b for b in backtest if b.get('bet_placed')])} paris back-test")

    wb = Workbook()
    wb.remove(wb.active)  # Supprimer feuille vide par défaut

    print("[Excel] Création Dashboard...")
    create_dashboard(wb, live_bets, backtest)

    print("[Excel] Création feuille Value Bets Live...")
    create_live_sheet(wb, live_bets)

    if backtest:
        print("[Excel] Création feuille Back-test...")
        create_backtest_sheet(wb, backtest)
        print("[Excel] Création feuille P&L Cumulé...")
        create_pnl_sheet(wb, backtest)

    wb.save(OUTPUT_FILE)
    print(f"\nFichier Excel genere : {OUTPUT_FILE}")
    print(f"   Ouvre-le dans Excel pour voir les résultats !")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="NBA Value Bot — Excel Tracker")
    parser.add_argument("--no-backtest", action="store_true",
                        help="Ne pas inclure le back-test")
    args = parser.parse_args()
    generate_excel(include_backtest=not args.no_backtest)
